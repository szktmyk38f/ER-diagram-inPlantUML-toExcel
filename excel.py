import re
import openpyxl
from openpyxl.utils import get_column_letter


wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = 'Sheet1'


test_data = open("test.pu", "r", encoding="utf-8_sig")
lines = test_data.readlines()

height = 0
is_first = True
seq = 1

for line in lines:
    # create header
    if is_first:
        ws1['C7'] = '#'
        ws1['D7'] = 'PK'
        ws1['E7'] = 'FK'
        ws1['F7'] = 'column name'
        ws1['G7'] = 'item name'
        ws1['H7'] = 'type'
        ws1['I7'] = 'length'
        ws1['J7'] = 'not null'
        ws1['K7'] = 'default'
        is_first = False

    if '}' in line:
        height += 3
        is_first = True
        continue

    if '--' in line or '@startuml' in line or '@enduml' in line:
        continue

    entity_index = line.find('entity')
    primary_index = line.find('PK')
    foreign_index = line.find('FK')

    if entity_index != -1:
        table_physical_name = re.search(r'entity "(.+)" ', line).group(1)
        table_logical_name = re.search(r'as (.+) {', line).group(1)
        ws1['A3'] = "physical name"
        ws1['A4'] = table_physical_name
        ws1['B3'] = "logical name"
        ws1['B4'] = table_logical_name
    elif primary_index != -1:
        ws1['C8'] = seq
        seq += 1
        if '+' in line:
            primary_key_name = re.search(r'\+ (.+) ', line).group(1)
            ws1['D8'] = '○'
            ws1['F8'] = primary_key_name
        elif '*' in line:
            primary_key_name = re.search(r'\* (.+) ', line).group(1)
            ws1['D8'] = '○'
            ws1['F8'] = primary_key_name
        else:
            primary_key_name = re.search(r'        (.+) ', line).group(1)
            ws1['D8'] = '○'
            ws1['F8'] = primary_key_name
    elif foreign_index != -1:
        ws1['C' + str(9 + height)] = seq
        seq += 1
        if '#' in line:
            foreign_key_name = re.search(r'# (.+) ', line).group(1)
            ws1['E' + str(9 + height)] = '○'
            ws1['F' + str(9 + height)] = foreign_key_name
        else:
            foreign_key_name = re.search(r'        (.+) ', line).group(1)
            ws1['E' + str(9 + height)] = '○'
            ws1['F' + str(9 + height)] = foreign_key_name
        height += 1
    else:
        print(line)
        line = line.strip()
        if len(line) != 0:
            ws1['C' + str(9 + height)] = seq
            ws1['F' + str(9 + height)] = line
            seq += 1
            height += 1


# set column width
for col in ws1.columns:
    max_length = 0
    column = col[0].column

    for cell in col:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))

    adjusted_width = (max_length + 2) * 1.2
    ws1.column_dimensions[get_column_letter(column)].width = adjusted_width

wb.save('test.xlsx')


test_data.close()
