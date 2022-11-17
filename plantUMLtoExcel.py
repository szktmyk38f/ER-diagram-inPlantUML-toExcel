import re
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side

side = Side(style='thin', color='000000')
border = Border(top=side, bottom=side, left=side, right=side)

wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = 'Sheet1'


test_data = open("test.pu", "r", encoding="utf-8_sig")
lines = test_data.readlines()

height = 0
is_first = True
seq = 1
start_ruled_line = 0
end_ruled_line = 0

for line in lines:
    # create header
    if is_first and '{' in line and 'entity' in line:
        ws1['C' + str(7 + height)] = '#'
        ws1['D' + str(7 + height)] = 'PK'
        ws1['E' + str(7 + height)] = 'FK'
        ws1['F' + str(7 + height)] = 'column name'
        ws1['G' + str(7 + height)] = 'item name'
        ws1['H' + str(7 + height)] = 'type'
        ws1['I' + str(7 + height)] = 'length'
        ws1['J' + str(7 + height)] = 'not null'
        ws1['K' + str(7 + height)] = 'default'
        is_first = False

    if '}' in line:
        end_ruled_line = height
        for row in ws1['C' + str(7 + start_ruled_line) + ':K' + str(8 + end_ruled_line)]:
            for cell in row:
                cell.border = border
        height += 7
        start_ruled_line = height
        is_first = True
        seq = 1
        continue

    if '--' in line or '@startuml' in line or '@enduml' in line or \
            '}|' in line or '}o' in line or '||' in line or '|{' in line or 'o{' in line or '||' in line:
        continue

    entity_index = line.find('entity')
    primary_index = line.find('PK')
    foreign_index = line.find('FK')

    if entity_index != -1:
        table_physical_name = re.search(r'entity "(.+)" ', line).group(1)
        table_logical_name = re.search(r'as (.+) {', line).group(1)
        ws1['A' + str(3 + height)] = "physical name"
        ws1['A' + str(4 + height)] = table_physical_name
        ws1['B' + str(3 + height)] = "logical name"
        ws1['B' + str(4 + height)] = table_logical_name
    elif primary_index != -1:
        ws1['C' + str(8 + height)] = seq
        seq += 1
        if '+' in line:
            primary_key_name = re.search(r'\+ (.+) ', line).group(1)
            ws1['D' + str(8 + height)] = '○'
            ws1['F' + str(8 + height)] = primary_key_name
        elif '*' in line:
            primary_key_name = re.search(r'\* (.+) ', line).group(1)
            ws1['D' + str(8 + height)] = '○'
            ws1['F' + str(8 + height)] = primary_key_name
        else:
            primary_key_name = re.search(r'        (.+) ', line).group(1)
            ws1['D' + str(8 + height)] = '○'
            ws1['F' + str(8 + height)] = primary_key_name
    elif foreign_index != -1:
        ws1['C' + str(9 + height)] = seq
        seq += 1
        if '#' in line:
            foreign_key_name = re.search(r'# (.+) ', line).group(1)
            ws1['E' + str(9 + height)] = '○'
            ws1['F' + str(9 + height)] = foreign_key_name
        else:
            foreign_key_name = re.search(r'        (.+) ', line).group(1)
            ws1['E' + str(9 + height)] = '○'
            ws1['F' + str(9 + height)] = foreign_key_name
        height += 1
    else:
        line = line.strip()
        if len(line) != 0:
            ws1['C' + str(9 + height)] = seq
            ws1['F' + str(9 + height)] = line
            seq += 1
            height += 1


# set column width
for col in ws1.columns:
    max_length = 0
    column = col[0].column

    for cell in col:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))

    adjusted_width = (max_length + 2) * 1.2
    ws1.column_dimensions[get_column_letter(column)].width = adjusted_width

wb.save('test.xlsx')


test_data.close()
